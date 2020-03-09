# Pyhton 下用Django语法增删查改，注意创建空的数据库需要在终端或数据库软件中，表结构需要在models.py模块
import os, django,datetime
from dateutil.relativedelta import relativedelta
os.environ.setdefault ( "DJANGO_SETTINGS_MODULE" , "IBS.settings" )  # project_name 项目名称


from collector import models
from django.contrib import auth
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
import pymysql
from openpyxl import load_workbook, Workbook

#后续开发：用户登录，封装成一个函数，调用传入用户名和密码即可


# 从EXCEL中导入数据,封装成一个函数，调用是传入表格路径即可,未来可用该函数导入预算，可用该函数导入不同的sheet数
def xlstodb(localpath):
    wb = load_workbook(localpath)
    ws = wb['Cover']
    mon = ws['C5'].value
    print(mon)
    Last_month = datetime.date.today () - relativedelta ( months=1 )
    if int(mon) == datetime.datetime.now().month or int(mon) == Last_month.month : #只能上传本月或上月报表
        ws = wb['ActualData']
        rows = ws.max_row
        columns = ws.max_column
        # 读取excel第一行的列名，写入list
        column_heading = [ws.cell(row=1, column=x).value for x in range(1, columns + 1)]
        column_name = ['amount', 'accountid', 'version']  # 数据库必需字段
        if len([name for name in column_name if name not in column_heading]) == 0:  # 返回字段组成的list为空，则说明文件列标题包含MySQL需要的字段
            print(' - 检查完成，执行写入')
            amt = column_heading.index('amount')
            aid = column_heading.index('accountid')
            vid = column_heading.index('version')
            # 判断从第几行开始取数
            if ws.cell(row=3,column=2).value == None:
                table_start_line = 4
            else :
                table_start_line = 3
            row = table_start_line
            rowdata = [] #存储单行数据
            rowlist = [] #存储多行数据
            for row in range(table_start_line, rows+1):
                for column in range(1, columns + 1):
                    rowdata.append(str(ws.cell(row=row, column=column).value))
                print(rowdata)
                print(rowdata[amt],rowdata[aid],rowdata[vid])
                rowlist.append(models.ActualData(amount=rowdata[amt],accountid_id=rowdata[aid],version_id=rowdata[vid]))
                rowdata = []
            # print('rowlist',rowlist)
            try :
                models.ActualData.objects.bulk_create(rowlist)  # 使用bulk_create批量导入
                msg = '导入成功'
            except Exception as e:
                print('异常', e)
                msg = '导入失败'
        else:
            print('文件列标题不完全包含数据库需要的字段，请检查文件。')
            msg = '文件列标题不完全包含数据库需要的字段，请检查文件。'
        wb.close()  # 关闭excel
        return msg
    else :
        print ( '此工作簿"封面"会计期间信息为%s月，系统要求只能上传本月或上月报表！执行导入程序前，退出后直接用本工作簿执行生成程序！' % (mon) )



import tkinter as tk
application_window = tk.Tk()
# Build a list of tuples for each file type the file dialog should display
my_filetypes = [('all files', '.*'), ('xlsx files', '.xlsx')]
from tkinter import filedialog
localpath = filedialog.askopenfilename(parent = application_window,initialdir = os.getcwd(),title="Please select a file:",filetypes = my_filetypes)

xlstodb(localpath)






