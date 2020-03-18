import time
import sys
import inspect
import pymysql
from collector import models, models_view
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl import load_workbook, Workbook
import os
import django
import datetime
from dateutil.relativedelta import relativedelta
os.environ.setdefault("DJANGO_SETTINGS_MODULE","IBS.settings")  # project_name 项目名称
django.setup()


# #Django models之related_name 和 _set,不可同时并存，但均适用于"一查询多"
# # 通过币种查询适用该币种的所有公司
# cny = models.Currency.objects.get(currencyname__contains='CNY')
# # print(cny.company_set.all()) #一旦在Foreykey中定义了related_name，此项命令即失效
# print(cny.comp_cur.all())
#
# # 没有定义related_name，则可用：（通过科目类型查询属于该科目类型的所有科目）
# accountgroup = models.AccountType.objects.get(typename='费用')
# print(accountgroup.account_set.all())


# 从models到数据库，从类到表和视图，再到excel的sheet和field，思路历程：
# print(models)
# print(dir(models))
# print(models.__name__)
# print(models.AccountBg.__str__)
# print(dir(models.AccountBg))
# print(models.AccountBg.__dict__)
# print(models.AccountBg.__name__)
# #取出表的中文名
# print(models.AccountBg._meta.verbose_name_plural)
#
# #取出字段的中文名
# fielddic1={}
# for field1 in models.AccountBg._meta.fields:
#     fielddic1[field1.name] = field1.verbose_name
# print(fielddic1)
#
# for name, obj in inspect.getmembers(sys.modules[__name__]):
#     if inspect.isclass(obj):
#         print(obj)
#         print(name)
#
# # 获取openpyxl类中所有方法列表：
# method_list1 = [func for func in dir(Protection) if callable(getattr(Protection, func))]
# print(len(method_list1))
# # 排除内置方法列表：
# method_list2 = [func for func in dir(Protection) if callable(getattr(Protection, func)) and not func.startswith("__")]
# print(method_list2)
#
# i = 0
# for method in inspect.getmembers(Protection):
#     print(method)
#     i = i+1
#     print(i)




def xlstodb(localpath, sheetlist):
    talblelist = []
    fieldlist = []
    wb = load_workbook(localpath)
    ws = wb['Cover']
    mon = ws['C5'].value
    print(mon)
    Last_month = datetime.date.today() - relativedelta(months=1)
    if int(mon) == datetime.datetime.now().month or int(
            mon) == Last_month.month:  # 只能上传本月或上月报表
        for stname, obj in inspect.getmembers(models):
            if inspect.isclass(obj) and stname in sheetlist:
                talblelist.append(stname)
                print(stname)
                for fieldobj in obj._meta.fields:
                    fieldlist.append(fieldobj.__dict__['name'])
                print(fieldlist)  # 临时存放某一张表的所有字段
                fieldlist = []
                ws = wb[stname]
                rowdata = []  # 存储单行数据
                rowlist = []  # 存储多行数据
                rows = ws.max_row
                columns = ws.max_column
                for row in range(3, rows + 1):
                    for column in range(1, columns + 1):
                        rowdata.append(
                            str(ws.cell(row=row, column=column).value))
                    print(rowdata)
                    rowlist.append(rowdata)
                    rowdata = []
                print(rowlist)
                try:
                    obj.bulk_create(rowlist)  # 使用bulk_create批量导入
                    msg = '导入成功'
                except Exception as e:
                    print ('异常', e)
                    msg = '导入失败'
            else:
                print('文件列标题不完全包含数据库需要的字段，请检查文件!')
                msg = '文件列标题不完全包含数据库需要的字段，请检查文件。'
            wb.close()  # 关闭excel
            return msg
    else:
        print (
            '此工作簿"封面"会计期间信息为%s月，系统要求只能上传本月或上月报表！执行导入程序前，退出后直接用本工作簿执行生成程序！' %
            (mon))
import tkinter as tk
application_window = tk.Tk()
application_window.withdraw() #隐藏消息框
# Build a list of tuples for each file type the file dialog should display
my_filetypes = [('xlsx files', '.xlsx')]
from tkinter import filedialog
localpath = filedialog.askopenfilename(parent = application_window,initialdir = os.getcwd(),title="请选择上传数据文件:",filetypes = my_filetypes)
sheetlist = ['ActualData', 'BudgetData', 'YtdData']
xlstodb(localpath,sheetlist)