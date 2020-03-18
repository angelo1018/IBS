import os, django,datetime,time,inspect
from dateutil.relativedelta import relativedelta
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "IBS.settings")  # project_name 项目名称
django.setup()
from collector import models,models_view
from django.contrib import auth
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
import pymysql
from openpyxl import load_workbook , Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import warnings
warnings.filterwarnings("ignore")

import tkinter as tk
application_window = tk.Tk()
application_window.withdraw() #隐藏消息框
# Build a list of tuples for each file type the file dialog should display
my_filetypes = [('xlsx files', '.xlsx')]
from tkinter import filedialog


# 从db中获取所有纳入填报范围的公司编码和公司名称，验正用户录入正确信息
rowlist = models.Company.objects.all().filter(isvalid=1)
comids = []
for rowobj in rowlist :
    # print(rowobj.serializable_value('companyname') + ':' + rowobj.serializable_value('companycode'))
    comids.append(rowobj.serializable_value('companycode'))
print(comids)
comid = input('请输入您的公司代码：').strip()
while comid not in comids :
    comid = input('请参照上述列表输入正确的公司代码：')


# 后续可封装成日期不同格式录入的函数检查（1、%Y-%m; 2、%Y-%m-%d 3、%Y-%m-%d %H:%M:%M）*************************


# 从DB中导出数据到EXCEL,封装成一个函数，调用是传入表格路径即可
def xltoxl(localpath):
    try:
        wb = load_workbook(localpath)
        ws = wb['Cover']
        ws['A4'].font = Font ( name='Arial', size=14 , color=colors.RED , italic=True )
        ws['B4'] = models.Company.objects.get(companycode=comid ).companyname
        ws['B5'] = valid_date.year
        ws['C5'] = valid_date.month
        ws['B6'] = str(models.Company.objects.get(companycode=comid).currency)
        vname = 'V' + comid + '-' + str ( valid_date.date()) + '/' + str ( dtoday.month + dtoday.day + dtoday.second )
        ws['B7'] = vname
        print('你本次生成的填报文件版本号为：', vname)
        # 版本号写入EXCEL同时写入后台数据库,同时作为后续可否导入数据的检查
        models.Version.objects.update_or_create(accountdate=valid_date, actualorbudget=True, versionname=vname,
                                        companycode_id=comid)

        # if comid == '0001' : #分析用户导出所有数据用，除总部外，其他单位只导出本单位数据*********************************
        #     for obj in models_view.MysqlView.objects.all().values_list():
        #         ws = wb['MysqlView']
        #         print(obj)
        #         ws.append(obj)
        # else:
        #     for obj in models_view.MysqlView.objects.all().values_list(companycode = comid):
        #         ws = wb['MysqlView']
        #         print ( obj )
        #         ws.append ( obj )
        #     ws.sheet_state = 'hidden'
        #     ws.protection.enable()
        #     ws.protection.set_password('sheet123')

        # 导出表新增内容
        excludest = ['ActualData','BudgetData','User'] #排除列表
        excludest =[]
        for stname, obj in inspect.getmembers(models):
            if inspect.isclass(obj):
                ws = wb[stname]
                maxrowinxl = ws.max_row
                maxrowindb = obj.objects.all().count()
                if stname not in excludest and maxrowindb > maxrowinxl - 2:
                    diffs = obj.objects.all()[maxrowinxl-1:]
                    for rows in diffs.values_list():
                        ws.append(list(rows))
                ws.sheet_state = 'hidden'
                ws.protection.enable()
                ws.protection.set_password('sheet123')
        for stname, obj in inspect.getmembers(models_view):
            if inspect.isclass(obj):
                ws = wb[stname]
                maxrowinxl = ws.max_row
                maxrowindb = obj.objects.all().count()
                if stname not in excludest and maxrowindb > maxrowinxl - 2:
                    diffs = obj.objects.all()[maxrowinxl-1:]
                    for rows in diffs.values_list():
                        ws.append(list(rows))
                ws.sheet_state = 'hidden'
                ws.protection.enable()
                ws.protection.set_password('sheet123')

        for stname in excludest:
            ws = wb[stname]
            ws.sheet_state = 'hidden'
            ws.protection.enable()
            ws.protection.set_password('sheet123')

        ws = wb['Cover']
        # ws.sheet_state = 'hidden'
        ws.protection.enable()
        ws.protection.set_password('sheet123')
        filename = 'FRP'+comid+ strdate+'.xlsx'
        wb.save(filename)
        print('成功生成填报文件，文件保存位置为:'+ os.getcwd())
    except Exception as e :
        print("异常错误",e)
        msg = '导出失败'


# 从界面录入正确的年和月# 检查日期格式是否正确
strdate = input('请输入当前或上一会计期间（格式为YYYY-MM）：')
dtoday = datetime.datetime.today()
Last_month = dtoday - relativedelta(months=1)
# 把输入的会计年月加上日

def is_valid_date(strdate):
    try:
        # datetime.datetime.strptime(strdate, "%Y-%m-%d")
        datetime.datetime.strptime(strdate, "%Y-%m")
        return True
    except:
        return False

while True:
    try:
        valid_ym = datetime.datetime.strptime ( strdate, "%Y-%m" )
        valid_date = dtoday.replace( year=valid_ym.year, month=valid_ym.month)
        # 只能生成本月或上月填报模板
        if is_valid_date(strdate) and (valid_ym.month == datetime.datetime.now().month or valid_ym.month == Last_month.month):
            localpath = filedialog.askopenfilename(parent=application_window, initialdir=os.getcwd(),
                                                     title="请选择模板生成本月可填报文件:" , filetypes=my_filetypes )
            xltoxl(localpath)
            break
    except Exception as e:
        print("异常错误", e)
    strdate = input('输入有误，请重新输入（正确输入举例：2020-03）：')







