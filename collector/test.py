import os, django,datetime
from dateutil.relativedelta import relativedelta
os.environ.setdefault ( "DJANGO_SETTINGS_MODULE" , "IBS.settings" )  # project_name 项目名称
django.setup()
#
# from collector import models
# from django.contrib import auth
# from django.contrib.auth import authenticate, login
# from django.contrib.auth.decorators import login_required
# import pymysql
# from openpyxl import load_workbook, Workbook

import os, django, pymysql, inspect, sys, datetime, time
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "IBS.settings")  # project_name 项目名称
django.setup()

from openpyxl import load_workbook , Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,colors

from collector import models,models_view


#Django models之related_name 和 _set,不可同时并存，但均适用于"一查询多"
# 通过币种查询适用该币种的所有公司
cny = models.Currency.objects.get(currencyname__contains='CNY')
# print(cny.company_set.all()) #一旦在Foreykey中定义了related_name，此项命令即失效
print(cny.comp_cur.all())

# 没有定义related_name，则可用：（通过科目类型查询属于该科目类型的所有科目）
accountgroup = models.AccountType.objects.get(typename='费用')
print(accountgroup.account_set.all())